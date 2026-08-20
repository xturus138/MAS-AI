import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook

def parse_numbered_list(text: str) -> list[str]:
    if not text:
        return []
    
    parts = re.split(r'\n*\d+\.\s+', text)
    cleaned = [p.strip() for p in parts if p.strip()]
    
    if not cleaned:
        cleaned = [p.strip() for p in text.split('\n') if p.strip()]
        
    return cleaned

def load_scenarios(xlsx_path: str) -> list[dict]:
    source_workbook = Path(xlsx_path).expanduser().resolve()
    workbook_fingerprint = hashlib.sha256(source_workbook.read_bytes()).hexdigest()
    wb = load_workbook(source_workbook, data_only=False)
    ws = wb.active
    
    scenarios = []
    header_found = False
    
    for source_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not row[0]:
            continue
            
        first_col = str(row[0]).strip()
        
        if not header_found:
            if first_col.upper() == "TCS ID":
                header_found = True
            continue
            
        tcs_id = first_col
        if not tcs_id or tcs_id.isdigit() or tcs_id.startswith("="):
            continue
                
        menu = str(row[1]).strip() if row[1] else "-"
        sub1 = str(row[2]).strip() if row[2] else "-"
        sub2 = str(row[3]).strip() if row[3] else "-"
        
        nav_parts = []
        if menu != "-": nav_parts.append(menu)
        if sub1 != "-": nav_parts.append(sub1)
        if sub2 != "-": nav_parts.append(sub2)
        nav_context = " > ".join(nav_parts) if nav_parts else "N/A"
        
        scenario_desc = str(row[4]).strip() if row[4] else ""
        raw_test_step = row[5] if len(row) > 5 else None
        test_step_raw = str(raw_test_step).strip() if raw_test_step else ""
        expected_raw = str(row[6]).strip() if row[6] else ""
        test_type = str(row[7]).strip() if len(row) > 7 and row[7] else "Pos."
        user_role = str(row[8]).strip() if len(row) > 8 and row[8] else "User"
        
        sub_steps = parse_numbered_list(test_step_raw)
        
        if sub_steps:
            scenarios.append({
                "tcs_id": tcs_id,
                "navigation_context": nav_context,
                "scenario_desc": scenario_desc,
                "test_type": test_type,
                "user_role": user_role,
                "sub_steps": sub_steps,
                "expected_result": expected_raw,
                "source_row": source_row,
                "raw_test_step": raw_test_step,
                "source_workbook": str(source_workbook),
                "workbook_fingerprint": workbook_fingerprint,
            })
            
    return scenarios
