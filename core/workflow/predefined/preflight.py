"""Read-only preflight checks for the Firebase Chat predefined baseline.

The validators in this module are pure with respect to devices and external
services.  Live device inspection is isolated behind :class:`ADBPreflightProbe`
and is never invoked by the validators themselves.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook

from core.utils.xlsx_loader import parse_numbered_list


FIREBASE_PACKAGE = "com.mikirinkode.firebasechatapp"
FIREBASE_CHAT_CASE_COUNT = 69
FIREBASE_CHAT_WORKBOOK_FINGERPRINT = (
    "937b45c789d397e34ce47c19e525383d2891e2787fecf20a40eac09e45de67d8"
)
FIREBASE_CHAT_ORIGINAL_HEADERS = (
    "TCS ID",
    "Menu",
    "Submenu 1",
    "Submenu 2",
    "Test Case Scenario",
    "Test Step",
    "Expected Result",
    "Test Type",
    "User",
    "Time Testing",
    "Testing Status",
    "Updated At",
    "Testing By",
    "OK Evid.",
    "Issue Status",
    "Updated At",
    "Scope",
    "Issue",
    "Developer Status",
    "Updated At",
    "PIC Now",
    "BE PIC",
    "FE PIC",
    "BE Target",
    "FE Target",
    "BE Notes",
    "FE Notes",
)
AGENT_ROLES = ("OBSERVER", "DECIDER", "REFLECTOR", "ORCHESTRATOR")
NO_CREDENTIAL_PROVIDERS = frozenset({"local", "9router"})


@dataclass(frozen=True)
class PreflightFinding:
    """One non-secret preflight result."""

    component: str
    code: str
    severity: str
    message: str
    details: dict[str, Any] = field(default_factory=dict)


@dataclass
class PreflightReport:
    """Structured, serializable findings from one or more preflight checks."""

    findings: list[PreflightFinding] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def blocking(self) -> bool:
        return any(finding.severity == "error" for finding in self.findings)

    @property
    def ready(self) -> bool:
        return not self.blocking

    def add(
        self,
        component: str,
        code: str,
        severity: str,
        message: str,
        **details: Any,
    ) -> None:
        self.findings.append(
            PreflightFinding(
                component=component,
                code=code,
                severity=severity,
                message=message,
                details=details,
            )
        )

    def merge(self, other: "PreflightReport") -> None:
        self.findings.extend(other.findings)
        self.metadata.update(other.metadata)

    def to_dict(self) -> dict[str, Any]:
        return {
            "ready": self.ready,
            "blocking": self.blocking,
            "findings": [asdict(finding) for finding in self.findings],
            "metadata": self.metadata,
        }


@dataclass(frozen=True)
class DeviceSnapshot:
    """Read-only device facts supplied to the pure device validator."""

    serial: str
    reachable: bool
    online: bool
    unlocked: bool
    firebase_package_installed: bool
    firebase_package_launchable: bool
    current_package: str
    current_activity: str
    declared_permissions_ready: bool | None
    permission_readiness_description: str
    account_ready: bool | None
    account_readiness_description: str
    installed_packages: tuple[str, ...] = ()


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _case_id(value: object) -> str | None:
    candidate = _clean(value)
    if not candidate or candidate.isdigit() or candidate.startswith("="):
        return None
    return candidate


def validate_workbook(
    workbook_path: str | Path,
    *,
    expected_fingerprint: str = FIREBASE_CHAT_WORKBOOK_FINGERPRINT,
    strict_firebase_chat: bool = True,
) -> PreflightReport:
    """Validate a scenario workbook.

    When ``strict_firebase_chat`` is ``True`` (the default, preserving prior
    behavior), the immutable Firebase Chat baseline contract is enforced:
    the pinned SHA-256 fingerprint, the full original 27-column header
    schema, and exactly 69 unique TCS IDs.

    When ``False``, only generic structural checks apply (readable
    workbook, a ``TCS ID`` header row, no duplicate TCS IDs, and no empty
    test steps), so other scenario workbooks can be run without being held
    to the Firebase Chat baseline's exact shape.
    """

    report = PreflightReport(metadata={"workbook": {}})
    metadata = report.metadata["workbook"]
    path = Path(workbook_path).expanduser().resolve()
    metadata["path"] = str(path)
    metadata["strict_firebase_chat"] = strict_firebase_chat

    if not path.is_file():
        metadata["file_state"] = "missing"
        report.add(
            "workbook",
            "workbook.missing",
            "error",
            "The Firebase Chat workbook does not exist.",
        )
        return report

    try:
        workbook_bytes = path.read_bytes()
        actual_fingerprint = hashlib.sha256(workbook_bytes).hexdigest()
    except OSError as exc:
        metadata["file_state"] = "unreadable"
        report.add(
            "workbook",
            "workbook.unreadable",
            "error",
            "The Firebase Chat workbook cannot be read.",
            error_type=type(exc).__name__,
        )
        return report

    metadata.update(
        file_state="readable",
        fingerprint=actual_fingerprint,
        expected_fingerprint=expected_fingerprint.lower(),
    )
    if strict_firebase_chat and actual_fingerprint.lower() != expected_fingerprint.lower():
        report.add(
            "workbook",
            "workbook.fingerprint",
            "error",
            "The workbook fingerprint does not match the approved Firebase Chat baseline.",
            actual=actual_fingerprint,
            expected=expected_fingerprint.lower(),
        )

    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook.active
        header_row = None
        headers: list[object] = []
        rows = list(sheet.iter_rows(values_only=True))
        for row_number, row in enumerate(rows, start=1):
            if row and _clean(row[0]).upper() == "TCS ID":
                header_row = row_number
                headers = list(row)
                break

        if header_row is None:
            metadata["header_state"] = "missing"
            report.add(
                "workbook",
                "workbook.headers",
                "error",
                "The original Firebase Chat header row is missing.",
            )
            return report

        normalized_headers = tuple(_clean(value) for value in headers)
        metadata["header_row"] = header_row
        if strict_firebase_chat:
            metadata["header_state"] = (
                "valid"
                if normalized_headers == FIREBASE_CHAT_ORIGINAL_HEADERS
                else "invalid"
            )
        else:
            metadata["header_state"] = "valid"
        updated_at_columns = [
            index
            for index, value in enumerate(normalized_headers, start=1)
            if value == "Updated At"
        ]
        metadata["first_updated_at_column"] = (
            updated_at_columns[0] if updated_at_columns else None
        )
        metadata["updated_at_columns"] = updated_at_columns

        if strict_firebase_chat and normalized_headers != FIREBASE_CHAT_ORIGINAL_HEADERS:
            report.add(
                "workbook",
                "workbook.headers",
                "error",
                "The workbook does not preserve the required original header schema.",
                expected=list(FIREBASE_CHAT_ORIGINAL_HEADERS),
                actual=list(normalized_headers),
            )

        test_step_column = (
            normalized_headers.index("Test Step")
            if "Test Step" in normalized_headers
            else None
        )
        case_ids: list[str] = []
        empty_steps: list[dict[str, object]] = []
        for row_number, row in enumerate(rows[header_row:], start=header_row + 1):
            if not row:
                continue
            tcs_id = _case_id(row[0])
            if tcs_id is None:
                continue
            case_ids.append(tcs_id)
            raw_step = (
                row[test_step_column]
                if test_step_column is not None and len(row) > test_step_column
                else None
            )
            raw_text = _clean(raw_step)
            parsed_steps = parse_numbered_list(raw_text)
            if not raw_text or not parsed_steps:
                empty_steps.append({"tcs_id": tcs_id, "source_row": row_number})

        normalized_ids = [case_id.casefold() for case_id in case_ids]
        unique_count = len(set(normalized_ids))
        metadata["case_count"] = len(case_ids)
        metadata["unique_tcs_id_count"] = unique_count

        if strict_firebase_chat:
            if len(case_ids) != FIREBASE_CHAT_CASE_COUNT or unique_count != FIREBASE_CHAT_CASE_COUNT:
                report.add(
                    "workbook",
                    "workbook.case_count",
                    "error",
                    "The Firebase Chat baseline must contain exactly 69 unique TCS IDs.",
                    case_count=len(case_ids),
                    unique_tcs_id_count=unique_count,
                    expected=FIREBASE_CHAT_CASE_COUNT,
                )
        elif not case_ids:
            report.add(
                "workbook",
                "workbook.case_count",
                "error",
                "The workbook does not contain any valid test cases.",
                case_count=0,
            )

        duplicates = sorted(
            {
                case_ids[index]
                for index, normalized in enumerate(normalized_ids)
                if normalized_ids.count(normalized) > 1
            }
        )
        if duplicates:
            report.add(
                "workbook",
                "workbook.duplicate_tcs_id",
                "error",
                "Duplicate TCS IDs are not allowed in the Firebase Chat baseline.",
                tcs_ids=duplicates,
            )

        if empty_steps:
            report.add(
                "workbook",
                "workbook.empty_test_step",
                "error",
                "Every Firebase Chat case must have nonempty raw and parsed test steps.",
                cases=empty_steps,
            )
    except Exception as exc:
        metadata["file_state"] = "unreadable"
        report.add(
            "workbook",
            "workbook.unreadable",
            "error",
            "The Firebase Chat workbook cannot be opened as an Excel workbook.",
            error_type=type(exc).__name__,
        )
    finally:
        if workbook is not None:
            workbook.close()

    return report


def _credential_is_present(config: Mapping[str, object], role: str, provider: str) -> bool:
    candidates = [f"{role}_API_KEY"]
    if provider in {"gemini", "google"}:
        candidates.extend(("GEMINI_API_KEY", "GOOGLE_API_KEY"))
    elif provider == "azure":
        candidates.append("AZURE_OPENAI_KEY")
    elif provider == "vertex":
        candidates.extend(("VERTEX_API_KEY", "GOOGLE_APPLICATION_CREDENTIALS"))
    else:
        candidates.append(f"{provider.upper().replace('-', '_')}_API_KEY")

    for key in candidates:
        value = _clean(config.get(key))
        if value and value.casefold() not in {"none", "not-needed-for-local"}:
            return True
    return False


def validate_config(config: Mapping[str, object]) -> PreflightReport:
    """Validate baseline configuration without importing or invoking DSE code."""

    report = PreflightReport(metadata={"config": {"agents": {}}})
    metadata = report.metadata["config"]
    mode = _clean(config.get("WORKFLOW_STRATEGY")).lower()
    metadata["workflow_mode"] = mode or "missing"
    if mode != "predefined":
        report.add(
            "config",
            "config.workflow_mode",
            "error",
            "The Firebase Chat baseline requires predefined workflow mode.",
            state="missing" if not mode else "invalid",
        )

    target_device = _clean(config.get("TARGET_DEVICE"))
    metadata["target_device_state"] = "configured" if target_device else "missing"
    if not target_device:
        report.add(
            "config",
            "config.target_device",
            "error",
            "A nonempty target device must be configured.",
            state="missing",
        )

    dse_value = config.get("OBSERVER_UNCERTAINTY_ENABLED")
    metadata["dse_state"] = "disabled" if dse_value is False else "enabled_or_invalid"
    if dse_value is not False:
        report.add(
            "config",
            "config.dse_enabled",
            "error",
            "Observer uncertainty must be explicitly False for the predefined baseline.",
            state="enabled_or_invalid",
        )

    for role in AGENT_ROLES:
        agent_name = role.lower()
        provider = _clean(config.get(f"{role}_PROVIDER")).lower()
        model = _clean(config.get(f"{role}_MODEL"))
        agent_metadata = {
            "provider_state": "configured" if provider else "missing",
            "model_state": "configured" if model else "missing",
        }
        metadata["agents"][agent_name] = agent_metadata

        if not provider:
            agent_metadata["credential_state"] = "not_checked"
            report.add(
                "config",
                "config.provider_missing",
                "error",
                f"{role.title()} provider is missing.",
                agent=agent_name,
                state="missing",
            )
        if not model:
            report.add(
                "config",
                "config.model_missing",
                "error",
                f"{role.title()} model is missing.",
                agent=agent_name,
                state="missing",
            )

        if not provider:
            continue
        if provider in NO_CREDENTIAL_PROVIDERS:
            agent_metadata["credential_state"] = "not_required"
        elif _credential_is_present(config, role, provider):
            agent_metadata["credential_state"] = "configured"
        else:
            agent_metadata["credential_state"] = "missing"
            report.add(
                "config",
                "config.credential_missing",
                "error",
                f"{role.title()} credentials are missing for the configured provider.",
                agent=agent_name,
                provider=provider,
                state="missing",
            )

    return report


def validate_device(
    snapshot: DeviceSnapshot,
    *,
    expected_serial: str = "",
    strict_firebase_chat: bool = True,
) -> PreflightReport:
    """Validate an injected snapshot without performing a device query.

    When ``strict_firebase_chat`` is ``True`` (the default), the Firebase
    Chat package must be installed and launchable or the report blocks.
    When ``False`` (any non-Firebase-Chat scenario), the Firebase Chat
    package's presence is only reported as informational metadata --
    ``snapshot.installed_packages`` already holds every package ``pm list
    packages`` returned, so callers are not limited to checking one
    hardcoded package.
    """

    account_state = (
        "unknown"
        if snapshot.account_ready is None
        else "ready"
        if snapshot.account_ready
        else "not_ready"
    )
    permission_state = (
        "unknown"
        if snapshot.declared_permissions_ready is None
        else "ready"
        if snapshot.declared_permissions_ready
        else "not_ready"
    )
    metadata = {
        "serial": snapshot.serial,
        "reachable": snapshot.reachable,
        "online": snapshot.online,
        "unlocked": snapshot.unlocked,
        "package": FIREBASE_PACKAGE,
        "package_installed": snapshot.firebase_package_installed,
        "package_launchable": snapshot.firebase_package_launchable,
        "package_check_enforced": strict_firebase_chat,
        "installed_package_count": len(snapshot.installed_packages),
        "current_package": snapshot.current_package,
        "current_activity": snapshot.current_activity,
        "permission_state": permission_state,
        "permission_readiness_description": snapshot.permission_readiness_description,
        "account_state": account_state,
        "account_readiness_description": snapshot.account_readiness_description,
    }
    report = PreflightReport(metadata={"device": metadata})

    if not snapshot.serial:
        report.add(
            "device",
            "device.serial_missing",
            "error",
            "The device snapshot has no serial.",
        )
    elif expected_serial and snapshot.serial != expected_serial:
        report.add(
            "device",
            "device.serial_mismatch",
            "error",
            "The snapshot serial does not match the configured target device.",
            expected=expected_serial,
            actual=snapshot.serial,
        )
    if not snapshot.reachable:
        report.add(
            "device",
            "device.unreachable",
            "error",
            "The target device is not reachable.",
        )
    if not snapshot.online:
        report.add(
            "device", "device.offline", "error", "The target device is offline."
        )
    if not snapshot.unlocked:
        report.add(
            "device", "device.locked", "error", "The target device is locked."
        )
    if not snapshot.firebase_package_installed:
        report.add(
            "device",
            "device.package_missing",
            "error" if strict_firebase_chat else "warning",
            "The Firebase Chat package is not installed."
            if strict_firebase_chat
            else "The Firebase Chat package is not installed (informational; "
            "this scenario does not require it).",
            package=FIREBASE_PACKAGE,
        )
    if not snapshot.firebase_package_launchable:
        report.add(
            "device",
            "device.package_not_launchable",
            "error" if strict_firebase_chat else "warning",
            "The Firebase Chat package has no launchable activity."
            if strict_firebase_chat
            else "The Firebase Chat package has no launchable activity "
            "(informational; this scenario does not require it).",
            package=FIREBASE_PACKAGE,
        )
    if snapshot.declared_permissions_ready is False:
        report.add(
            "device",
            "device.permissions_not_ready",
            "error",
            "Declared runtime permissions are not ready.",
            state="not_ready",
        )
    elif snapshot.declared_permissions_ready is None:
        report.add(
            "device",
            "device.permissions_unknown",
            "warning",
            "Declared runtime permission readiness could not be checked.",
            state="unknown",
        )
    if snapshot.account_ready is False:
        report.add(
            "device",
            "device.account_not_ready",
            "error",
            "The declared Firebase account setup is not ready.",
            state="not_ready",
        )
    elif snapshot.account_ready is None:
        report.add(
            "device",
            "device.account_unknown",
            "warning",
            "Firebase account readiness could not be checked automatically.",
            state="unknown",
        )

    return report


def run_preflight(
    *,
    workbook_path: str | Path,
    config: Mapping[str, object],
    device_snapshot: DeviceSnapshot,
    expected_fingerprint: str = FIREBASE_CHAT_WORKBOOK_FINGERPRINT,
    strict_firebase_chat: bool = True,
) -> PreflightReport:
    """Combine pure workbook, config, and injected-snapshot validation."""

    report = PreflightReport()
    report.merge(
        validate_workbook(
            workbook_path,
            expected_fingerprint=expected_fingerprint,
            strict_firebase_chat=strict_firebase_chat,
        )
    )
    report.merge(validate_config(config))
    report.merge(
        validate_device(
            device_snapshot,
            expected_serial=_clean(config.get("TARGET_DEVICE")),
            strict_firebase_chat=strict_firebase_chat,
        )
    )
    return report


def write_preflight_report(report: PreflightReport, destination: str | Path) -> Path:
    """Atomically write a secret-free JSON representation of ``report``."""

    path = Path(destination).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            json.dump(report.to_dict(), handle, indent=2, ensure_ascii=False)
            handle.write("\n")
            temporary_path = Path(handle.name)
        os.replace(temporary_path, path)
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()
    return path


class ADBPreflightProbe:
    """Live adapter that performs direct, read-only ADB shell queries only.

    Importing or constructing this class does not create a device handle.
    ``capture`` is the sole live boundary and deliberately performs no server
    setup, app start/stop, input, permission mutation, install, reset, or
    data-clearing operation.
    """

    def __init__(
        self,
        serial: str,
        *,
        adb_client: object | None = None,
        declared_permissions: Sequence[str] = (),
        account_readiness_description: str = (
            "Account readiness is not checkable through read-only ADB queries."
        ),
    ) -> None:
        self.serial = serial
        self.adb_client = adb_client
        self.declared_permissions = tuple(declared_permissions)
        self.account_readiness_description = account_readiness_description

    @staticmethod
    def _shell(device: object, command: str) -> tuple[str, int | None]:
        if hasattr(device, "shell2"):
            result = device.shell2(command)
            return (
                _clean(getattr(result, "output", "")),
                getattr(result, "returncode", None),
            )
        result = device.shell(command)
        return _clean(getattr(result, "output", result)), None

    @staticmethod
    def _resolved_firebase_component(output: str, returncode: int | None) -> bool:
        if returncode != 0:
            return False
        for line in output.splitlines():
            match = re.fullmatch(
                rf"{re.escape(FIREBASE_PACKAGE)}/([A-Za-z0-9_.$]+)",
                line.strip(),
            )
            if not match:
                continue
            activity = match.group(1)
            if activity.startswith(".") or activity.startswith(
                f"{FIREBASE_PACKAGE}."
            ):
                return True
        return False

    @staticmethod
    def _current_component(activity_dump: str) -> tuple[str, str]:
        for line in activity_dump.splitlines():
            if "ResumedActivity" not in line:
                continue
            match = re.search(r"\b([A-Za-z0-9_.]+)/([A-Za-z0-9_.$]+)\b", line)
            if match:
                return match.group(1), match.group(2)
        return "", ""

    def capture(self) -> DeviceSnapshot:
        """Create an ADB handle, issue read-only queries, and return facts."""

        try:
            if self.adb_client is None:
                from adbutils import adb

                client = adb
            else:
                client = self.adb_client
            device = client.device(self.serial)
            boot_completed, boot_returncode = self._shell(
                device, "getprop sys.boot_completed"
            )
            power_state, power_returncode = self._shell(device, "dumpsys power")
            window_policy, policy_returncode = self._shell(
                device, "dumpsys window policy"
            )
            package_path, package_returncode = self._shell(
                device, f"pm path {FIREBASE_PACKAGE}"
            )
            launchable, launch_returncode = self._shell(
                device,
                f"cmd package resolve-activity --brief {FIREBASE_PACKAGE}",
            )
            activity_dump, _ = self._shell(device, "dumpsys activity activities")
            current_package, current_activity = self._current_component(activity_dump)
            package_dump, package_dump_returncode = self._shell(
                device, f"dumpsys package {FIREBASE_PACKAGE}"
            )
            packages_output, packages_returncode = self._shell(
                device, "pm list packages"
            )
            installed_packages = tuple(
                sorted(
                    line.strip()[len("package:") :]
                    for line in packages_output.splitlines()
                    if line.strip().startswith("package:")
                )
                if packages_returncode == 0
                else ()
            )

            lock_markers = (
                "mShowingLockscreen=true",
                "isStatusBarKeyguard=true",
                "mDreamingLockscreen=true",
            )
            screen_on = power_returncode == 0 and any(
                marker in power_state
                for marker in ("mWakefulness=Awake", "state=ON")
            )
            unlocked = (
                screen_on
                and policy_returncode == 0
                and not any(marker in window_policy for marker in lock_markers)
            )
            package_installed = package_returncode == 0 and any(
                line.strip().startswith("package:") for line in package_path.splitlines()
            )
            package_launchable = self._resolved_firebase_component(
                launchable, launch_returncode
            )

            if not self.declared_permissions:
                permissions_ready = None
                permission_description = (
                    "No required permission list was supplied for read-only validation."
                )
            elif package_dump_returncode != 0:
                permissions_ready = None
                permission_description = (
                    "Declared runtime permission readiness could not be queried."
                )
            else:
                missing_permissions = [
                    permission
                    for permission in self.declared_permissions
                    if not re.search(
                        rf"{re.escape(permission)}[^\n]*granted=true",
                        package_dump,
                    )
                ]
                permissions_ready = not missing_permissions
                permission_description = (
                    "All declared runtime permissions are granted."
                    if permissions_ready
                    else "One or more declared runtime permissions are not granted."
                )

            return DeviceSnapshot(
                serial=self.serial,
                reachable=True,
                online=boot_returncode == 0 and boot_completed == "1",
                unlocked=unlocked,
                firebase_package_installed=package_installed,
                firebase_package_launchable=package_launchable,
                current_package=current_package,
                current_activity=current_activity,
                declared_permissions_ready=permissions_ready,
                permission_readiness_description=permission_description,
                account_ready=None,
                account_readiness_description=self.account_readiness_description,
                installed_packages=installed_packages,
            )
        except Exception as exc:
            return DeviceSnapshot(
                serial=self.serial,
                reachable=False,
                online=False,
                unlocked=False,
                firebase_package_installed=False,
                firebase_package_launchable=False,
                current_package="",
                current_activity="",
                declared_permissions_ready=None,
                permission_readiness_description=(
                    f"Device permission readiness is unavailable ({type(exc).__name__})."
                ),
                account_ready=None,
                account_readiness_description=self.account_readiness_description,
            )
