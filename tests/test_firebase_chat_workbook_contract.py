"""Workbook-boundary tests for the Firebase Chat predefined baseline."""

from __future__ import annotations

import importlib
import shutil
import sys
import types
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

from agents.recorder_agent import RecorderAgent
from core.utils.xlsx_loader import load_scenarios
from core.utils.output_manager import create_run_output


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = REPO_ROOT / "scenarios" / "firebase_chat" / "scenario.xlsx"
ALLOWED_EXECUTION_HEADERS = {
    "Time Testing",
    "Testing Status",
    "Updated At",
    "Testing By",
    "OK Evid.",
    "Issue Status",
}


def _header_row_and_values(ws):
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] and str(row[0]).strip().upper() == "TCS ID":
            return row_index, list(row)
    raise AssertionError("Firebase Chat workbook must contain a TCS ID header row")


def _target_row(ws, tcs_id: str) -> int:
    header_row, _ = _header_row_and_values(ws)
    for row_index in range(header_row + 1, ws.max_row + 1):
        if str(ws.cell(row_index, 1).value).strip() == tcs_id:
            return row_index
    raise AssertionError(f"{tcs_id} must exist in the copied workbook")


def _report_metrics(status: str = "functional_anomaly") -> dict:
    return {
        "status": status,
        "total_duration_seconds": 65,
        "timestamp": "2026-08-14 15:00:00",
        "mode": "predefined",
        "justification": {"reflector_final_judgment": "Unexpected result"},
    }


def _write_report(tmp_path: Path, source_workbook: Path = SOURCE_WORKBOOK):
    evidence_dir = tmp_path / "case-evidence"
    state = {
        "tcs_id": "FC-LGN-001",
        "output_dir": str(evidence_dir),
        "reports_dir": str(tmp_path / "reports"),
    }
    RecorderAgent().write_test_report(
        state,
        _report_metrics(),
        source_workbook=str(source_workbook),
    )
    return tmp_path / "reports" / "test_report.xlsx"


def _report_state(tmp_path: Path, tcs_id: str, name: str) -> dict:
    return {
        "tcs_id": tcs_id,
        "output_dir": str(tmp_path / f"{name}-evidence"),
        "reports_dir": str(tmp_path / name / "reports"),
    }


def test_firebase_chat_loader_keeps_ordered_raw_workbook_metadata():
    scenarios = load_scenarios(str(SOURCE_WORKBOOK))
    reloaded_scenarios = load_scenarios(str(SOURCE_WORKBOOK))
    source = openpyxl.load_workbook(SOURCE_WORKBOOK, data_only=False)
    sheet = source.active

    assert len(scenarios) == 69
    assert [scenario["tcs_id"] for scenario in scenarios] == [
        sheet.cell(row, 1).value
        for row in range(1, sheet.max_row + 1)
        if isinstance(sheet.cell(row, 1).value, str)
        and sheet.cell(row, 1).value.startswith("FC-")
    ]
    assert scenarios[0]["tcs_id"] == "FC-LGN-001"
    assert scenarios[-1]["tcs_id"] == "FC-PRF-004"

    fingerprints = {scenario["workbook_fingerprint"] for scenario in scenarios}
    assert len(fingerprints) == 1
    assert len(next(iter(fingerprints))) == 64
    assert [scenario["workbook_fingerprint"] for scenario in reloaded_scenarios] == [
        scenario["workbook_fingerprint"] for scenario in scenarios
    ]

    for scenario in scenarios:
        assert scenario["source_workbook"] == str(SOURCE_WORKBOOK.resolve())
        assert isinstance(scenario["source_row"], int)
        assert scenario["raw_test_step"] == sheet.cell(scenario["source_row"], 6).value
        assert scenario["raw_test_step"]


def test_report_preserves_duplicate_headers_and_changes_only_six_target_cells(tmp_path):
    source_copy = tmp_path / "source.xlsx"
    shutil.copy2(SOURCE_WORKBOOK, source_copy)
    before = openpyxl.load_workbook(source_copy, data_only=False)
    supplementary_sheet = before.create_sheet("Supplementary")
    supplementary_sheet["B2"] = "must remain untouched"
    supplementary_sheet["B2"].font = Font(bold=True, color="00AA00")
    before.save(source_copy)
    before = openpyxl.load_workbook(source_copy, data_only=False)
    before_headers = _header_row_and_values(before.active)[1]

    report_path = _write_report(tmp_path, source_copy)
    after = openpyxl.load_workbook(report_path, data_only=False)

    assert after.sheetnames == before.sheetnames
    assert after["Supplementary"]["B2"].value == "must remain untouched"
    assert after["Supplementary"]["B2"].font.bold is True
    assert after["Supplementary"]["B2"].font.color.rgb == before["Supplementary"]["B2"].font.color.rgb
    assert _header_row_and_values(after.active)[1] == before_headers
    assert _header_row_and_values(after.active)[1].count("Updated At") == 3

    target_row = _target_row(before.active, "FC-LGN-001")
    first_updated_at_column = before_headers.index("Updated At") + 1
    allowed_columns = {
        column_index
        for column_index, header in enumerate(before_headers, start=1)
        if header in ALLOWED_EXECUTION_HEADERS
        and not (header == "Updated At" and column_index != first_updated_at_column)
    }
    changed_cells = set()
    for sheet_index, before_sheet in enumerate(before.worksheets):
        after_sheet = after.worksheets[sheet_index]
        assert after_sheet.max_row == before_sheet.max_row
        assert after_sheet.max_column == before_sheet.max_column
        for row_index in range(1, before_sheet.max_row + 1):
            for column_index in range(1, before_sheet.max_column + 1):
                before_value = before_sheet.cell(row_index, column_index).value
                after_value = after_sheet.cell(row_index, column_index).value
                if before_value != after_value:
                    changed_cells.add((sheet_index, row_index, column_index))

    assert changed_cells == {
        (0, target_row, column_index) for column_index in allowed_columns
    }
    assert after.active.cell(target_row, first_updated_at_column).value == "2026-08-14 15:00:00"
    for column_index, header in enumerate(before_headers, start=1):
        if header == "Updated At" and column_index != first_updated_at_column:
            assert after.active.cell(target_row, column_index).value == before.active.cell(
                target_row, column_index
            ).value


def test_report_uses_explicit_source_path_outside_current_working_directory(tmp_path, monkeypatch):
    source_copy = tmp_path / "explicit-source.xlsx"
    unrelated_directory = tmp_path / "unrelated-current-directory"
    unrelated_directory.mkdir()
    shutil.copy2(SOURCE_WORKBOOK, source_copy)
    monkeypatch.chdir(unrelated_directory)

    report_path = _write_report(tmp_path, source_copy)

    assert report_path.exists()
    assert openpyxl.load_workbook(report_path, data_only=False).active["A13"].value == "FC-LGN-001"


@pytest.mark.parametrize(
    ("status", "testing_status", "issue_status"),
    [
        ("success", "OK", "OK"),
        ("functional_anomaly", "NG", "Functional anomaly"),
        ("stagnated", "NG", "Stagnated"),
        ("technical_error", "NG", "Technical error"),
    ],
)
def test_report_maps_canonical_outcomes_without_calling_an_anomaly_a_bug(
    tmp_path, status, testing_status, issue_status
):
    source_copy = tmp_path / f"{status}-source.xlsx"
    shutil.copy2(SOURCE_WORKBOOK, source_copy)
    evidence_dir = tmp_path / f"{status}-evidence"
    state = {
        "tcs_id": "FC-LGN-001",
        "output_dir": str(evidence_dir),
        "reports_dir": str(tmp_path / status / "reports"),
    }

    RecorderAgent().write_test_report(
        state,
        _report_metrics(status),
        source_workbook=str(source_copy),
    )

    report = openpyxl.load_workbook(
        tmp_path / status / "reports" / "test_report.xlsx", data_only=False
    )
    target_row = _target_row(report.active, "FC-LGN-001")
    _, headers = _header_row_and_values(report.active)
    assert report.active.cell(target_row, headers.index("Testing Status") + 1).value == testing_status
    actual_issue_status = report.active.cell(target_row, headers.index("Issue Status") + 1).value
    assert actual_issue_status == issue_status
    assert "bug" not in actual_issue_status.lower()


def test_report_rejects_missing_execution_header_before_creating_destination(tmp_path):
    source_copy = tmp_path / "missing-header.xlsx"
    shutil.copy2(SOURCE_WORKBOOK, source_copy)
    workbook = openpyxl.load_workbook(source_copy)
    header_row, headers = _header_row_and_values(workbook.active)
    workbook.active.cell(header_row, headers.index("Issue Status") + 1).value = ""
    workbook.save(source_copy)

    evidence_dir = tmp_path / "case-evidence"
    destination = tmp_path / "reports" / "test_report.xlsx"
    state = {
        "tcs_id": "FC-LGN-001",
        "output_dir": str(evidence_dir),
        "reports_dir": str(destination.parent),
    }

    with pytest.raises(ValueError, match="Issue Status"):
        RecorderAgent().write_test_report(
            state,
            _report_metrics(),
            source_workbook=str(source_copy),
        )

    assert not destination.exists()


def test_shared_report_preserves_the_first_case_when_the_second_case_is_added(tmp_path):
    source_copy = tmp_path / "source.xlsx"
    shared_dir = tmp_path / "shared"
    shared_dir.mkdir()
    shutil.copy2(SOURCE_WORKBOOK, source_copy)
    recorder = RecorderAgent()
    first_state = _report_state(tmp_path, "FC-LGN-001", "first")
    second_state = _report_state(tmp_path, "FC-LGN-002", "second")

    recorder.write_test_report(
        first_state,
        _report_metrics("success"),
        source_workbook=str(source_copy),
        shared_dir=str(shared_dir),
    )
    recorder.write_test_report(
        second_state,
        _report_metrics("functional_anomaly"),
        source_workbook=str(source_copy),
        shared_dir=str(shared_dir),
    )

    shared_report = openpyxl.load_workbook(shared_dir / "test_report.xlsx", data_only=False)
    header_row, headers = _header_row_and_values(shared_report.active)
    first_row = _target_row(shared_report.active, "FC-LGN-001")
    second_row = _target_row(shared_report.active, "FC-LGN-002")
    assert first_row > header_row
    assert shared_report.active.cell(first_row, headers.index("Testing Status") + 1).value == "OK"
    assert shared_report.active.cell(second_row, headers.index("Testing Status") + 1).value == "NG"
    assert shared_report.active.cell(first_row, headers.index("OK Evid.") + 1).value == first_state["output_dir"]
    assert shared_report.active.cell(second_row, headers.index("OK Evid.") + 1).value == second_state["output_dir"]


def test_shared_report_keeps_existing_file_when_a_second_case_write_fails(tmp_path):
    source_copy = tmp_path / "source.xlsx"
    shared_dir = tmp_path / "shared"
    shared_dir.mkdir()
    shutil.copy2(SOURCE_WORKBOOK, source_copy)
    first_state = _report_state(tmp_path, "FC-LGN-001", "first")
    second_state = _report_state(tmp_path, "FC-LGN-002", "second")
    RecorderAgent().write_test_report(
        first_state,
        _report_metrics("success"),
        source_workbook=str(source_copy),
        shared_dir=str(shared_dir),
    )

    class _FailAfterSharedWorkbookWrite(RecorderAgent):
        def _fill_report_sheet(self, dest, tcs_id, output_dir, metrics):
            super()._fill_report_sheet(dest, tcs_id, output_dir, metrics)
            if Path(dest).parent == shared_dir:
                raise OSError("simulated shared-report write failure")

    with pytest.raises(OSError, match="simulated shared-report write failure"):
        _FailAfterSharedWorkbookWrite().write_test_report(
            second_state,
            _report_metrics("functional_anomaly"),
            source_workbook=str(source_copy),
            shared_dir=str(shared_dir),
        )

    shared_report = openpyxl.load_workbook(shared_dir / "test_report.xlsx", data_only=False)
    _, headers = _header_row_and_values(shared_report.active)
    first_row = _target_row(shared_report.active, "FC-LGN-001")
    second_row = _target_row(shared_report.active, "FC-LGN-002")
    assert shared_report.active.cell(first_row, headers.index("Testing Status") + 1).value == "OK"
    assert shared_report.active.cell(second_row, headers.index("Testing Status") + 1).value is None


def test_report_closes_workbook_when_saving_fails(tmp_path, monkeypatch):
    report_copy = tmp_path / "report.xlsx"
    shutil.copy2(SOURCE_WORKBOOK, report_copy)
    real_workbook = openpyxl.load_workbook(report_copy)

    class _SaveFailingWorkbook:
        def __init__(self):
            self.worksheets = real_workbook.worksheets
            self.closed = False

        def save(self, _dest):
            raise OSError("simulated save failure")

        def close(self):
            self.closed = True
            real_workbook.close()

    wrapped_workbook = _SaveFailingWorkbook()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: wrapped_workbook)
    try:
        with pytest.raises(OSError, match="simulated save failure"):
            RecorderAgent()._fill_report_sheet(
                str(report_copy),
                "FC-LGN-001",
                str(tmp_path / "evidence"),
                _report_metrics(),
            )
        assert wrapped_workbook.closed is True
    finally:
        if not wrapped_workbook.closed:
            wrapped_workbook.close()


@pytest.mark.parametrize(
    ("runner_module_name", "entrypoint"),
    [
        ("core.workflow.predefined.runner", "run_predefined"),
        ("core.workflow.autonomous.runner", "run_autonomous"),
    ],
)
def test_runner_requires_explicit_workbook_instead_of_discovering_cwd_copy(
    tmp_path, monkeypatch, capsys, runner_module_name, entrypoint
):
    shutil.copy2(SOURCE_WORKBOOK, tmp_path / "scenario.xlsx")
    monkeypatch.chdir(tmp_path)
    real_easyocr = sys.modules.get("easyocr")
    sys.modules["easyocr"] = types.ModuleType("easyocr")
    try:
        runner_module = importlib.import_module(runner_module_name)
    finally:
        sys.modules.pop(runner_module_name, None)
        sys.modules.pop("tools.observer_tools", None)
        if real_easyocr is None:
            sys.modules.pop("easyocr", None)
        else:
            sys.modules["easyocr"] = real_easyocr

    monkeypatch.setattr(runner_module, "compute_run_root", lambda _mode: (tmp_path / "runs", "20260814", 1))
    monkeypatch.setattr(
        runner_module,
        "build_figma_adapter_from_prompt",
        lambda **_kwargs: None,
        raising=False,
    )

    class _UnexpectedDeviceAdapter:
        def __init__(self, *_args, **_kwargs):
            raise AssertionError("runner discovered a CWD workbook and reached device setup")

    monkeypatch.setattr(runner_module, "ADBAdapter", _UnexpectedDeviceAdapter, raising=False)

    getattr(runner_module, entrypoint)("")

    assert "Explicit scenario workbook path is required" in capsys.readouterr().out


def test_output_paths_can_use_coordinator_attempt_directory(tmp_path):
    attempt_dir = tmp_path / "run" / "scenario_01" / "attempt_003"

    paths = create_run_output(
        mode="predefined",
        tcs_id="FC-LGN-001",
        timestamp="20260814_120000",
        attempt_dir=str(attempt_dir),
    )

    assert Path(paths.run_dir) == attempt_dir
    assert Path(paths.steps_dir) == attempt_dir / "steps"
    assert Path(paths.logs_dir) == attempt_dir / "logs"
    assert Path(paths.reports_dir) == attempt_dir / "reports"
