from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from desktop_app.data.workbook import WorkbookCase, load_workbook_cases


def _write_minimal_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["TCS ID", "Menu", "Sub1", "Sub2", "Scenario", "Test Step", "Expected", "Type", "Role"])
    ws.append(["TCS-001", "Auth", "-", "-", "Login with valid credentials", "1. Tap email\n2. Tap password", "Home visible", "Pos.", "User"])
    ws.append(["TCS-002", "Chat", "-", "-", "Send text message", "1. Open chat\n2. Type\n3. Send", "Message sent", "Pos.", "User"])
    wb.save(path)


def test_load_workbook_cases_without_manifest_are_all_untested(tmp_path):
    xlsx_path = tmp_path / "scenario.xlsx"
    _write_minimal_workbook(xlsx_path)

    cases = load_workbook_cases(str(xlsx_path), manifest=None)

    assert len(cases) == 2
    assert cases[0] == WorkbookCase(
        tcs_id="TCS-001",
        module="Auth",
        scenario_title="Login with valid credentials",
        steps=2,
        precondition="Auth",
        status="untested",
    )
    assert cases[1].tcs_id == "TCS-002"
    assert cases[1].steps == 3
    assert cases[1].status == "untested"


def test_load_workbook_cases_joins_manifest_status(tmp_path):
    xlsx_path = tmp_path / "scenario.xlsx"
    _write_minimal_workbook(xlsx_path)
    manifest = {
        "cases": [
            {"tcs_id": "TCS-001", "status": "success"},
            {"tcs_id": "TCS-002", "status": "technical_error"},
        ]
    }

    cases = load_workbook_cases(str(xlsx_path), manifest=manifest)

    assert cases[0].status == "success"
    assert cases[1].status == "technical_error"


def test_load_workbook_cases_defaults_unmatched_case_to_untested(tmp_path):
    xlsx_path = tmp_path / "scenario.xlsx"
    _write_minimal_workbook(xlsx_path)
    manifest = {"cases": [{"tcs_id": "TCS-001", "status": "success"}]}

    cases = load_workbook_cases(str(xlsx_path), manifest=manifest)

    assert cases[0].status == "success"
    assert cases[1].status == "untested"
