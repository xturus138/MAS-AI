from __future__ import annotations

import importlib
import json
import subprocess
import sys
from contextlib import AbstractContextManager
from pathlib import Path

import openpyxl
import pytest

from core.workflow.predefined.preflight import (
    DeviceSnapshot,
    PreflightReport,
    validate_config,
)
from core.workflow.predefined.recovery import RecoveryCoordinator


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = REPO_ROOT / "scenarios" / "firebase_chat" / "scenario.xlsx"


def _ready_report(component: str) -> PreflightReport:
    return PreflightReport(metadata={component: {}})


def _blocking_report(component: str, code: str) -> PreflightReport:
    report = _ready_report(component)
    report.add(component, code, "error", f"Fake blocking {component} failure.")
    return report


def _valid_config(**overrides) -> dict[str, object]:
    values: dict[str, object] = {
        "WORKFLOW_STRATEGY": "predefined",
        "TARGET_DEVICE": "fake-device-001",
        "OBSERVER_UNCERTAINTY_ENABLED": False,
    }
    for role in ("OBSERVER", "DECIDER", "REFLECTOR", "ORCHESTRATOR"):
        values[f"{role}_PROVIDER"] = "local"
        values[f"{role}_MODEL"] = f"fake-{role.lower()}"
        values[f"{role}_API_KEY"] = ""
    values.update(overrides)
    return values


def _snapshot() -> DeviceSnapshot:
    return DeviceSnapshot(
        serial="fake-device-001",
        reachable=True,
        online=True,
        unlocked=True,
        firebase_package_installed=True,
        firebase_package_launchable=True,
        current_package="com.mikirinkode.firebasechatapp",
        current_activity=".MainActivity",
        declared_permissions_ready=True,
        permission_readiness_description="Fake permissions ready.",
        account_ready=True,
        account_readiness_description="Fake accounts ready.",
    )


def _scenario(tcs_id: str) -> dict:
    return {
        "tcs_id": tcs_id,
        "navigation_context": "Chat detail for Ada",
        "scenario_desc": f"Scenario {tcs_id}",
        "sub_steps": ["Open chat", "Send message"],
        "raw_test_step": "1. Open chat\n2. Send message",
        "workbook_fingerprint": "fake-fingerprint",
    }


class _Runtime(AbstractContextManager):
    def __init__(self, executor, events: list[str]):
        self.executor = executor
        self.events = events

    def __enter__(self):
        self.events.append("runtime_enter")
        return self.executor

    def __exit__(self, exc_type, exc, traceback):
        self.events.append("runtime_exit")
        return False


def _dependencies(
    tmp_path: Path,
    *,
    scenarios: list[dict],
    executor,
    events: list[str],
    config_values: dict[str, object] | None = None,
    workbook_report: PreflightReport | None = None,
    config_validator=validate_config,
    aggregate_report_writer=None,
):
    runner = importlib.import_module("core.workflow.predefined.runner")

    workbook_report = workbook_report or _ready_report("workbook")
    workbook_report.metadata.setdefault("workbook", {})["fingerprint"] = "fake-fingerprint"

    def probe_factory(serial):
        events.append("probe_constructed")

        class Probe:
            def capture(self):
                events.append("probe_captured")
                return _snapshot()

        return Probe()

    def runtime_factory(**kwargs):
        events.append("runtime_constructed")
        return _Runtime(executor, events)

    def run_root_factory(_mode):
        events.append("run_root_computed")
        return str(tmp_path / "run"), "2026-08-14", 1

    if aggregate_report_writer is None:
        aggregate_report_writer = lambda **kwargs: events.append("aggregate_written")

    return runner.RunnerDependencies(
        config_values=config_values or _valid_config(),
        workbook_validator=lambda _path: workbook_report,
        config_validator=config_validator,
        device_probe_factory=probe_factory,
        device_validator=lambda _snapshot, **_kwargs: _ready_report("device"),
        scenario_loader=lambda _path: scenarios,
        run_root_factory=run_root_factory,
        runtime_factory=runtime_factory,
        aggregate_report_writer=aggregate_report_writer,
    )


def test_workbook_or_config_failure_prevents_probe_runtime_and_run_root(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    events: list[str] = []
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")
    deps = _dependencies(
        tmp_path,
        scenarios=[_scenario("TCS-1")],
        executor=lambda *_args: {"status": "success"},
        events=events,
        workbook_report=_blocking_report("workbook", "workbook.fake"),
    )

    report = runner.run_predefined(str(workbook), dependencies=deps)

    assert report.blocking is True
    assert events == []


def test_dse_enabled_blocks_batch_and_fake_uncertainty_hook_is_never_called(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    events: list[str] = []
    uncertainty_calls: list[str] = []
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")
    deps = _dependencies(
        tmp_path,
        scenarios=[_scenario("TCS-1")],
        executor=lambda *_args: uncertainty_calls.append("called") or {"status": "success"},
        events=events,
        config_values=_valid_config(OBSERVER_UNCERTAINTY_ENABLED=True),
    )

    report = runner.run_predefined(str(workbook), dependencies=deps)

    assert report.blocking is True
    assert uncertainty_calls == []
    assert events == []


def test_explicit_predefined_mode_overrides_an_autonomous_environment_for_preflight(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    events: list[str] = []
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")
    deps = _dependencies(
        tmp_path,
        scenarios=[_scenario("TCS-1")],
        executor=lambda *_args: {"status": "success", "duration_seconds": 0},
        events=events,
        config_values=_valid_config(WORKFLOW_STRATEGY="autonomous"),
    )

    manifest = runner.run_predefined(str(workbook), dependencies=deps)

    assert manifest["cases"][0]["status"] == "success"
    assert "probe_captured" in events


def test_integrated_batch_continues_after_case_error_and_writes_mixed_aggregates(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    events: list[str] = []
    executed: list[str] = []
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")

    def execute(case, evidence_dir):
        executed.append(case["tcs_id"])
        if case["tcs_id"] == "TCS-1":
            raise RuntimeError("fake integrated failure")
        return {
            "status": "functional_anomaly",
            "duration_seconds": 2,
            "reason_ref": "expected did not match actual",
        }

    aggregate_calls = []

    def aggregate_report_writer(**kwargs):
        aggregate_calls.append(kwargs)
        (Path(kwargs["run_root"]) / "test_report.xlsx").write_bytes(b"fake excel")

    deps = _dependencies(
        tmp_path,
        scenarios=[_scenario("TCS-1"), _scenario("TCS-2")],
        executor=execute,
        events=events,
        aggregate_report_writer=aggregate_report_writer,
    )

    manifest = runner.run_predefined(str(workbook), dependencies=deps)

    assert executed == ["TCS-1", "TCS-2"]
    assert [case["status"] for case in manifest["cases"]] == [
        "technical_error",
        "functional_anomaly",
    ]
    assert (tmp_path / "run" / "batch_summary.json").is_file()
    assert (tmp_path / "run" / "batch_summary.csv").is_file()
    assert (tmp_path / "run" / "test_report.xlsx").is_file()
    assert aggregate_calls[0]["manifest"] is manifest


def test_integrated_resume_skips_terminal_and_allocates_new_interrupted_attempt(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")
    cases = [_scenario("TCS-1"), _scenario("TCS-2"), _scenario("TCS-3")]
    first_events: list[str] = []

    def interrupt_second(case, evidence_dir):
        if case["tcs_id"] == "TCS-2":
            (evidence_dir / "before-interrupt.txt").write_text("immutable", encoding="utf-8")
            raise KeyboardInterrupt()
        return {"status": "success", "duration_seconds": 1}

    with pytest.raises(KeyboardInterrupt):
        runner.run_predefined(
            str(workbook),
            dependencies=_dependencies(
                tmp_path,
                scenarios=cases,
                executor=interrupt_second,
                events=first_events,
            ),
        )

    resumed: list[tuple[str, str]] = []

    def resume_execute(case, evidence_dir):
        resumed.append((case["tcs_id"], evidence_dir.name))
        return {"status": "success", "duration_seconds": 1}

    manifest = runner.run_predefined(
        str(workbook),
        resume=str(tmp_path / "run"),
        dependencies=_dependencies(
            tmp_path,
            scenarios=cases,
            executor=resume_execute,
            events=[],
        ),
    )

    assert resumed == [("TCS-2", "attempt_002"), ("TCS-3", "attempt_001")]
    assert (tmp_path / "run" / "scenario_02" / "attempt_001" / "before-interrupt.txt").read_text(
        encoding="utf-8"
    ) == "immutable"
    assert [attempt["status"] for attempt in manifest["cases"][1]["attempts"]] == [
        "interrupted",
        "success",
    ]


def test_resume_requires_existing_manifest_before_probe_runtime_or_directory_creation(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    from core.workflow.predefined.batch import BatchCompatibilityError

    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")
    typo_root = tmp_path / "typo-resume-root"
    events: list[str] = []

    with pytest.raises(BatchCompatibilityError, match="batch_manifest.json"):
        runner.run_predefined(
            str(workbook),
            resume=str(typo_root),
            dependencies=_dependencies(
                tmp_path,
                scenarios=[_scenario("TCS-1")],
                executor=lambda *_args: {"status": "success"},
                events=events,
            ),
        )

    assert not typo_root.exists()
    assert events == []


def test_integrated_recovery_and_runner_never_modify_workbook_steps_or_add_bridge(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")
    first = _scenario("TCS-1")
    second = _scenario("TCS-2")
    originals = {
        case["tcs_id"]: (case["raw_test_step"], list(case["sub_steps"]))
        for case in (first, second)
    }

    def execute(case, evidence_dir):
        RecoveryCoordinator(
            observe_fresh=lambda _case: {"text": "Home"},
            assess_context=lambda _case, _observation: {"matched": False, "reason": "Wrong screen"},
            plan_transition=lambda _case, _observation, _assessment: ["Open chats"],
            execute_transition=lambda _case, _steps: {"artifact_path": "recovery.log"},
            verify_final=lambda _case, _observation, _execution: {"passed": True},
        ).run(case, evidence_dir)
        return {"status": "success", "duration_seconds": 0}

    runner.run_predefined(
        str(workbook),
        dependencies=_dependencies(
            tmp_path,
            scenarios=[first, second],
            executor=execute,
            events=[],
        ),
    )

    for case in (first, second):
        assert (case["raw_test_step"], case["sub_steps"]) == originals[case["tcs_id"]]


@pytest.mark.parametrize(
    ("state", "total_steps", "expected"),
    [
        ({"is_completed": True, "steps_completed_count": 2, "last_reflector_passed": True}, 2, "success"),
        ({"retry_exhausted": True, "is_completed": True, "stagnation_count": 0}, 1, "functional_anomaly"),
        ({"is_completed": False, "stagnation_count": 3}, 1, "stagnated"),
        ({"retry_exhausted": True, "stagnation_count": 3}, 1, "stagnated"),
        ({"technical_error_history": [{"reason": "fake"}], "retry_exhausted": True, "stagnation_count": 3}, 1, "technical_error"),
        ({"technical_error_history": [{"reason": "[SYSTEM_ERROR] fake"}]}, 1, "technical_error"),
    ],
)
def test_outcome_classification_distinguishes_all_four_statuses(state, total_steps, expected):
    runner = importlib.import_module("core.workflow.predefined.runner")

    outcome = runner.derive_case_outcome(state, total_steps=total_steps)

    assert outcome["status"] == expected


def test_predefined_graph_stops_once_stagnation_threshold_is_reached():
    from langgraph.graph import END

    from core.workflow.predefined.graph import _next_step_or_end

    assert _next_step_or_end({"is_completed": False, "stagnation_count": 3}) == END


def test_technical_fallback_has_precedence_over_assumed_completion():
    runner = importlib.import_module("core.workflow.predefined.runner")

    outcome = runner.derive_case_outcome(
        {
            "is_completed": True,
            "steps_completed_count": 2,
            "last_reflector_passed": True,
            "execution_result": "ERROR (Execution Failed): fake adb outage",
            "action_plan": {"reasoning": "LLM output parsing failed after recovery attempts"},
        },
        total_steps=2,
    )

    assert outcome["status"] == "technical_error"


def test_orchestrator_preserves_system_error_history_and_retry_exhaustion_is_not_success():
    from core.workflow.predefined.orchestrator import PredefinedOrchestrator

    class Procedural:
        def get_steps(self, _tcs_id, _entry_type):
            return ["Step one", "Step two"]

    class Memory:
        procedural = Procedural()

        def update(self, _packet):
            return None

    orchestrator = PredefinedOrchestrator(memory=Memory())
    system_error = orchestrator.orchestrate(
        {
            "tcs_id": "TCS-1",
            "sender": "reflector",
            "last_reflector_passed": False,
            "last_reflector_reasoning": "[SYSTEM_ERROR] fake reflector outage",
            "current_sub_step_index": 0,
            "current_step": 1,
            "step_retry_count": 0,
            "technical_error_history": [],
        }
    )
    retry_exhausted = orchestrator.orchestrate(
        {
            "tcs_id": "TCS-1",
            "sender": "reflector",
            "last_reflector_passed": False,
            "last_reflector_reasoning": "Expected text missing",
            "current_sub_step_index": 0,
            "current_step": 4,
            "step_retry_count": 3,
        }
    )

    assert system_error["technical_error_history"] == [
        {"step": 1, "reason": "[SYSTEM_ERROR] fake reflector outage"}
    ]
    assert system_error["is_completed"] is True
    assert retry_exhausted["retry_exhausted"] is True
    assert runner_status(retry_exhausted) == "functional_anomaly"


@pytest.mark.parametrize(
    "fallback_reason",
    [
        "Loading check error (assuming loaded): fake timeout",
        "UI change check returned None (assuming changed)",
    ],
)
def test_reflector_assumed_success_fallback_is_preserved_as_technical_history(
    tmp_path, fallback_reason
):
    from agents.reflector_agent import ReflectorAgent

    reflector = object.__new__(ReflectorAgent)
    reflector.memory = None
    reflector.logger = None
    update = reflector._build_return(
        state={"technical_error_history": []},
        passed=True,
        reasoning="Validity check passed",
        figma_discrepancies="",
        screenshot_path="",
        post_action_path=None,
        current_step=3,
        current_idx=0,
        figma_enabled=False,
        memory_context="",
        verification_chain={
            "loading_done": True,
            "loading_reasoning": fallback_reason,
            "ui_changed": True,
            "ui_change_reasoning": fallback_reason,
        },
    )

    assert update["technical_error_history"] == [
        {"step": 3, "reason": fallback_reason}
    ]


def test_executor_caught_exception_is_preserved_as_technical_history(monkeypatch):
    from agents.executor_agent import ExecutorAgent

    class Tools:
        d = None

        def click_coordinates(self, _x, _y):
            raise RuntimeError("fake adb outage")

    monkeypatch.setattr("agents.executor_agent.time.sleep", lambda _seconds: None)
    executor = ExecutorAgent(Tools())
    update = executor.execute(
        {
            "current_step": 7,
            "action_plan": {
                "action_type": "click",
                "target_id": 1,
                "intent": "Open chat",
                "is_completed": False,
            },
            "widgets": [{"id": 1, "bounds": [0, 0, 10, 10], "text": "Chat"}],
            "technical_error_history": [],
        }
    )

    assert update["technical_error_history"] == [
        {"step": 7, "reason": "ERROR (Execution Failed): fake adb outage"}
    ]


def test_decider_parse_fallback_is_preserved_as_technical_history(monkeypatch):
    from agents.decider_agent import ActionPlan, DeciderAgent

    class Prompt:
        def format_messages(self, **_kwargs):
            return []

    decider = object.__new__(DeciderAgent)
    decider.prompt = Prompt()
    decider.memory = None
    decider.logger = None
    decider.monitor = None
    monkeypatch.setattr(
        decider,
        "_invoke_with_recovery",
        lambda _messages, _step: ActionPlan(
            reasoning="LLM output parsing failed after recovery attempts: fake",
            action_type="none",
            intent="Fallback due to LLM output format error",
            target_id=-1,
            text_payload="",
            scroll_direction="",
            app_package="",
            is_completed=True,
        ),
    )
    update = decider.decide(
        {
            "current_step": 5,
            "orchestrator_instruction": "Open chat",
            "observer_analysis": "Home",
            "widgets": [],
            "technical_error_history": [],
        }
    )

    assert update["technical_error_history"] == [
        {
            "step": 5,
            "reason": "LLM output parsing failed after recovery attempts: fake",
        }
    ]


def runner_status(state: dict) -> str:
    runner = importlib.import_module("core.workflow.predefined.runner")
    return runner.derive_case_outcome(state, total_steps=1)["status"]


def test_aggregate_excel_is_rebuilt_from_manifest_including_skipped_terminal_cases(tmp_path):
    from agents.recorder_agent import RecorderAgent

    manifest = {
        "cases": [
            {
                "tcs_id": "FC-LGN-001",
                "status": "success",
                "duration_seconds": 1,
                "completed_at": "2026-08-14T01:02:03+00:00",
                "evidence_path": str(tmp_path / "attempt-old"),
            },
            {
                "tcs_id": "FC-LGN-002",
                "status": "technical_error",
                "duration_seconds": 2,
                "completed_at": "2026-08-14T02:03:04+00:00",
                "evidence_path": str(tmp_path / "attempt-new"),
            },
        ]
    }
    destination = tmp_path / "run" / "test_report.xlsx"

    RecorderAgent.write_batch_report(
        manifest=manifest,
        source_workbook=str(SOURCE_WORKBOOK),
        destination=str(destination),
    )

    workbook = openpyxl.load_workbook(destination, data_only=False)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in next(row for row in sheet.iter_rows() if row[0].value == "TCS ID")]
        rows = {
            str(sheet.cell(row, 1).value): row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row, 1).value
        }
        assert sheet.cell(rows["FC-LGN-001"], headers.index("Testing Status") + 1).value == "OK"
        assert sheet.cell(rows["FC-LGN-002"], headers.index("Issue Status") + 1).value == "Technical error"
        assert sheet.cell(rows["FC-LGN-001"], headers.index("OK Evid.") + 1).value == str(
            tmp_path / "attempt-old"
        )
        assert sheet.cell(rows["FC-LGN-001"], headers.index("Updated At") + 1).value == (
            "2026-08-14T01:02:03+00:00"
        )
    finally:
        workbook.close()


def test_cli_parser_exposes_preflight_and_resume_and_rejects_autonomous_resume(tmp_path):
    import main

    parser = main.build_parser()
    args = parser.parse_args(
        ["--scenario", "firebase_chat", "--mode", "predefined", "--preflight-only", "--resume", str(tmp_path)]
    )
    assert args.preflight_only is True
    assert args.resume == str(tmp_path)
    main.validate_cli_options(parser, args, effective_mode="predefined")

    autonomous = parser.parse_args(
        ["--scenario", "firebase_chat", "--mode", "autonomous", "--resume", str(tmp_path)]
    )
    with pytest.raises(SystemExit):
        main.validate_cli_options(parser, autonomous, effective_mode="autonomous")


def test_preflight_only_stops_before_run_root_runtime_and_aggregate(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")
    events: list[str] = []
    workbook = tmp_path / "scenario.xlsx"
    workbook.write_bytes(b"fake workbook boundary")

    report = runner.run_predefined(
        str(workbook),
        preflight_only=True,
        dependencies=_dependencies(
            tmp_path,
            scenarios=[_scenario("TCS-1")],
            executor=lambda *_args: {"status": "success"},
            events=events,
        ),
    )

    assert report.ready is True
    assert events == ["probe_constructed", "probe_captured"]


def test_missing_required_case_reporting_artifact_is_a_technical_failure(tmp_path):
    runner = importlib.import_module("core.workflow.predefined.runner")

    with pytest.raises(RuntimeError, match="reporting artifact"):
        runner.require_case_reporting_artifacts(tmp_path)

    (tmp_path / "final_metrics.json").write_text("{}", encoding="utf-8")
    reports = tmp_path / "reports"
    reports.mkdir()
    (reports / "test_report.xlsx").write_bytes(b"fake report")
    runner.require_case_reporting_artifacts(tmp_path)


def test_cleanup_failure_rewrites_final_metrics_and_case_excel_to_technical(tmp_path):
    from agents.recorder_agent import RecorderAgent

    runner = importlib.import_module("core.workflow.predefined.runner")
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    metrics = {
        "tcs_id": "FC-LGN-001",
        "status": "success",
        "timestamp": "2026-08-14T03:04:05+00:00",
        "mode": "predefined",
        "total_duration_seconds": 2,
    }
    (tmp_path / "final_metrics.json").write_text(json.dumps(metrics), encoding="utf-8")
    RecorderAgent().write_test_report(
        {
            "tcs_id": "FC-LGN-001",
            "output_dir": str(tmp_path),
            "reports_dir": str(reports_dir),
        },
        metrics,
        source_workbook=str(SOURCE_WORKBOOK),
    )

    rewritten = runner.reconcile_cleanup_failure_artifacts(
        attempt_dir=tmp_path,
        source_workbook=str(SOURCE_WORKBOOK),
        tcs_id="FC-LGN-001",
        metrics=metrics,
        cleanup_error=RuntimeError("fake memory close failure"),
    )

    assert rewritten["status"] == "technical_error"
    persisted = json.loads((tmp_path / "final_metrics.json").read_text(encoding="utf-8"))
    assert persisted["status"] == "technical_error"
    assert "fake memory close failure" in persisted["error_or_anomaly_reason"]
    workbook = openpyxl.load_workbook(reports_dir / "test_report.xlsx")
    try:
        sheet = workbook.active
        header_row = next(row for row in sheet.iter_rows() if row[0].value == "TCS ID")
        headers = [cell.value for cell in header_row]
        data_row = next(
            row for row in sheet.iter_rows() if str(row[0].value) == "FC-LGN-001"
        )
        assert data_row[headers.index("Testing Status")].value == "NG"
        assert data_row[headers.index("Issue Status")].value == "Technical error"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("initial_matched", "planned_steps", "final_matched"),
    [
        (False, [], True),
        (False, ["Open chats"], False),
        (False, ["Open chats"], None),
    ],
)
def test_recovery_validation_fails_closed_for_empty_plan_or_unverified_transition(
    initial_matched, planned_steps, final_matched
):
    runner = importlib.import_module("core.workflow.predefined.runner")

    with pytest.raises(RuntimeError, match="recovery"):
        runner.require_verified_recovery(
            initial_assessment={"matched": initial_matched},
            planned_steps=planned_steps,
            final_assessment={"matched": final_matched},
        )


@pytest.mark.parametrize(
    "technical_update",
    [
        {"technical_error_history": [{"reason": "fake decider parse fallback"}]},
        {"execution_result": "ERROR (Execution Failed): fake adb outage"},
    ],
)
def test_recovery_technical_agent_update_aborts_before_final_screen_verification(
    tmp_path, technical_update
):
    runner = importlib.import_module("core.workflow.predefined.runner")
    verification_calls: list[str] = []

    def execute_transition(_case, _steps):
        runner._agent_update(technical_update, phase="Recovery Executor")
        return {"steps_executed": 1}

    coordinator = RecoveryCoordinator(
        observe_fresh=lambda _case: {"text": "Home"},
        assess_context=lambda _case, _observation: {
            "matched": False,
            "reason": "Wrong screen",
        },
        plan_transition=lambda _case, _observation, _assessment: ["Open chats"],
        execute_transition=execute_transition,
        verify_final=lambda _case, _observation, _execution: verification_calls.append(
            "called"
        )
        or {"passed": True},
    )

    with pytest.raises(RuntimeError, match="Recovery Executor"):
        coordinator.run(_scenario("TCS-1"), tmp_path)

    assert verification_calls == []


def test_recovery_assessment_and_planning_use_navigation_text_without_figma():
    from types import SimpleNamespace

    from core.workflow.predefined.orchestrator import PredefinedOrchestrator

    class Structured:
        def __init__(self, model_name):
            self.model_name = model_name

        def invoke(self, messages):
            rendered = "\n".join(str(message.content) for message in messages)
            assert "Chat detail for Ada" in rendered
            assert "Home screen" in rendered
            if self.model_name == "NavigationContextAssessment":
                return SimpleNamespace(matched=False, reason="Chat detail is not open.")
            return SimpleNamespace(bridge_steps=["Open chats", "Open Ada"], reasoning="Minimal route")

    class FakeLLM:
        def with_structured_output(self, model):
            return Structured(model.__name__)

    orchestrator = PredefinedOrchestrator(llm=FakeLLM(), figma_adapter=None)

    assessment = orchestrator.assess_navigation_context(
        observation_text="Home screen",
        required_context="Chat detail for Ada",
    )
    steps = orchestrator.plan_recovery_transition(
        observation_text="Home screen",
        required_context="Chat detail for Ada",
        figma_context={},
    )

    assert assessment == {"matched": False, "reason": "Chat detail is not open."}
    assert steps == ["Open chats", "Open Ada"]


def test_observer_import_does_not_load_dse_service_clusterer_or_config_when_disabled():
    code = r'''
import importlib.abc
import sys

BLOCKED = {
    "core.uncertainty.clusterer",
    "core.uncertainty.config",
    "core.uncertainty.service",
}

class RejectDSEImplementation(importlib.abc.MetaPathFinder):
    def find_spec(self, fullname, path=None, target=None):
        if fullname in BLOCKED:
            raise AssertionError(f"DSE implementation import attempted: {fullname}")
        return None

sys.meta_path.insert(0, RejectDSEImplementation())
from agents.observer_agent import ObserverAgent
assert "core.uncertainty.service" not in sys.modules
'''
    completed = subprocess.run(
        [sys.executable, "-c", code],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
