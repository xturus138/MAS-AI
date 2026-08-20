"""Pure preflight tests for the 69-case Firebase Chat baseline."""

from __future__ import annotations

import importlib
import builtins
import json
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = REPO_ROOT / "scenarios" / "firebase_chat" / "scenario.xlsx"
EXPECTED_FINGERPRINT = (
    "937b45c789d397e34ce47c19e525383d2891e2787fecf20a40eac09e45de67d8"
)


def _preflight_module():
    try:
        return importlib.import_module("core.workflow.predefined.preflight")
    except ModuleNotFoundError:
        pytest.fail("Task 4 preflight module is not implemented")


def _copy_firebase_workbook(tmp_path: Path) -> Path:
    destination = tmp_path / "firebase-chat.xlsx"
    shutil.copy2(SOURCE_WORKBOOK, destination)
    return destination


def _header_row_and_values(workbook_path: Path) -> tuple[int, list[object]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        for row_number, row in enumerate(
            workbook.active.iter_rows(values_only=True), start=1
        ):
            if row and str(row[0]).strip().upper() == "TCS ID":
                return row_number, list(row)
    finally:
        workbook.close()
    raise AssertionError("fixture must contain a TCS ID header")


def _mutate_workbook(workbook_path: Path, mutation: str) -> None:
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        sheet = workbook.active
        header_row, headers = _header_row_and_values(workbook_path)
        case_rows = [
            row_number
            for row_number in range(header_row + 1, sheet.max_row + 1)
            if sheet.cell(row_number, 1).value
            and not str(sheet.cell(row_number, 1).value).strip().isdigit()
            and not str(sheet.cell(row_number, 1).value).strip().startswith("=")
        ]
        if mutation == "count":
            sheet.delete_rows(case_rows[-1])
        elif mutation == "header":
            sheet.cell(header_row, headers.index("Expected Result") + 1).value = (
                "Wrong Header"
            )
        elif mutation == "duplicate":
            sheet.cell(case_rows[1], 1).value = sheet.cell(case_rows[0], 1).value
        elif mutation == "empty_step":
            sheet.cell(case_rows[0], headers.index("Test Step") + 1).value = ""
        else:  # pragma: no cover - test helper guard
            raise AssertionError(f"unknown mutation: {mutation}")
        workbook.save(workbook_path)
    finally:
        workbook.close()


def _finding_codes(report, severity: str | None = None) -> set[str]:
    return {
        finding.code
        for finding in report.findings
        if severity is None or finding.severity == severity
    }


@pytest.fixture
def valid_config() -> dict[str, object]:
    config: dict[str, object] = {
        "WORKFLOW_STRATEGY": "predefined",
        "TARGET_DEVICE": "fake-device-001",
        "OBSERVER_UNCERTAINTY_ENABLED": False,
    }
    for role in ("OBSERVER", "DECIDER", "REFLECTOR", "ORCHESTRATOR"):
        config[f"{role}_PROVIDER"] = "openai"
        config[f"{role}_MODEL"] = f"fake-{role.lower()}-model"
        config[f"{role}_API_KEY"] = "super-secret-token"
    return config


def _valid_snapshot(module, *, account_ready: bool | None = True):
    return module.DeviceSnapshot(
        serial="fake-device-001",
        reachable=True,
        online=True,
        unlocked=True,
        firebase_package_installed=True,
        firebase_package_launchable=True,
        current_package="com.mikirinkode.firebasechatapp",
        current_activity=".MainActivity",
        declared_permissions_ready=True,
        permission_readiness_description="All declared runtime permissions are ready.",
        account_ready=account_ready,
        account_readiness_description=(
            "Fake Firebase accounts are ready."
            if account_ready is not None
            else "Account readiness cannot be checked by the fake snapshot."
        ),
    )


def test_temporary_copy_of_actual_firebase_workbook_passes_with_69_cases(tmp_path):
    module = _preflight_module()
    workbook_path = _copy_firebase_workbook(tmp_path)

    report = module.validate_workbook(
        workbook_path, expected_fingerprint=EXPECTED_FINGERPRINT
    )

    assert report.blocking is False
    assert report.metadata["workbook"]["case_count"] == 69
    assert report.metadata["workbook"]["unique_tcs_id_count"] == 69
    assert report.metadata["workbook"]["first_updated_at_column"] == 12


@pytest.mark.parametrize(
    ("mutation", "expected_code"),
    [
        pytest.param("count", "workbook.case_count", id="invalid-count"),
        pytest.param("header", "workbook.headers", id="invalid-header"),
        pytest.param("duplicate", "workbook.duplicate_tcs_id", id="duplicate-id"),
        pytest.param("empty_step", "workbook.empty_test_step", id="empty-step"),
    ],
)
def test_invalid_workbook_contract_is_blocking(tmp_path, mutation, expected_code):
    module = _preflight_module()
    workbook_path = _copy_firebase_workbook(tmp_path)
    _mutate_workbook(workbook_path, mutation)

    report = module.validate_workbook(
        workbook_path, expected_fingerprint=EXPECTED_FINGERPRINT
    )

    assert report.blocking is True
    assert expected_code in _finding_codes(report, "error")


@pytest.mark.parametrize("path_kind", ["missing", "corrupt"])
def test_missing_or_unreadable_workbook_is_blocking(tmp_path, path_kind):
    module = _preflight_module()
    workbook_path = tmp_path / "firebase-chat.xlsx"
    if path_kind == "corrupt":
        workbook_path.write_bytes(b"not an xlsx workbook")

    report = module.validate_workbook(
        workbook_path, expected_fingerprint=EXPECTED_FINGERPRINT
    )

    assert report.blocking is True
    assert _finding_codes(report, "error") & {
        "workbook.missing",
        "workbook.unreadable",
    }


def test_dse_true_blocks_baseline_and_strict_false_passes(valid_config):
    module = _preflight_module()
    enabled = dict(valid_config, OBSERVER_UNCERTAINTY_ENABLED=True)

    enabled_report = module.validate_config(enabled)
    disabled_report = module.validate_config(valid_config)

    assert "config.dse_enabled" in _finding_codes(enabled_report, "error")
    assert "config.dse_enabled" not in _finding_codes(disabled_report, "error")
    assert disabled_report.metadata["config"]["dse_state"] == "disabled"


def test_config_requires_predefined_target_models_providers_and_credentials(valid_config):
    module = _preflight_module()
    invalid = dict(valid_config)
    invalid.update(
        WORKFLOW_STRATEGY="autonomous",
        TARGET_DEVICE="",
        OBSERVER_PROVIDER="",
        DECIDER_MODEL="",
        REFLECTOR_API_KEY="",
    )

    report = module.validate_config(invalid)

    assert {
        "config.workflow_mode",
        "config.target_device",
        "config.provider_missing",
        "config.model_missing",
        "config.credential_missing",
    } <= _finding_codes(report, "error")


def test_local_provider_does_not_require_an_api_key(valid_config):
    module = _preflight_module()
    local = dict(valid_config, OBSERVER_PROVIDER="local", OBSERVER_API_KEY="")

    report = module.validate_config(local)

    observer = report.metadata["config"]["agents"]["observer"]
    assert observer["credential_state"] == "not_required"
    assert not any(
        finding.code == "config.credential_missing"
        and finding.details.get("agent") == "observer"
        for finding in report.findings
    )


def test_import_and_pure_config_validation_never_load_dse_service(tmp_path):
    code = r'''
import importlib.abc
import sys

class RejectDSE(importlib.abc.MetaPathFinder):
    def find_spec(self, fullname, path=None, target=None):
        if fullname.startswith("core.uncertainty"):
            raise AssertionError(f"DSE import attempted: {fullname}")
        return None

sys.meta_path.insert(0, RejectDSE())
from core.workflow.predefined.preflight import validate_config

config = {
    "WORKFLOW_STRATEGY": "predefined",
    "TARGET_DEVICE": "fake-device-001",
    "OBSERVER_UNCERTAINTY_ENABLED": False,
}
for role in ("OBSERVER", "DECIDER", "REFLECTOR", "ORCHESTRATOR"):
    config[f"{role}_PROVIDER"] = "local"
    config[f"{role}_MODEL"] = "fake-model"
    config[f"{role}_API_KEY"] = ""
assert not validate_config(config).blocking
assert "core.uncertainty.service" not in sys.modules
'''

    completed = subprocess.run(
        [sys.executable, "-c", code],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr


@pytest.mark.parametrize(
    ("changes", "expected_code"),
    [
        pytest.param({"online": False}, "device.offline", id="offline"),
        pytest.param(
            {"firebase_package_installed": False},
            "device.package_missing",
            id="missing-package",
        ),
        pytest.param({"unlocked": False}, "device.locked", id="locked"),
    ],
)
def test_fake_offline_missing_package_and_locked_device_are_blocking(
    changes, expected_code
):
    module = _preflight_module()
    values = _valid_snapshot(module).__dict__ | changes
    snapshot = module.DeviceSnapshot(**values)

    report = module.validate_device(snapshot, expected_serial="fake-device-001")

    assert report.blocking is True
    assert expected_code in _finding_codes(report, "error")


def test_unknown_account_readiness_is_visible_warning_not_success():
    module = _preflight_module()
    snapshot = _valid_snapshot(module, account_ready=None)

    report = module.validate_device(snapshot, expected_serial="fake-device-001")

    assert report.blocking is False
    assert "device.account_unknown" in _finding_codes(report, "warning")
    assert report.metadata["device"]["account_state"] == "unknown"


def test_json_report_has_configured_and_missing_states_but_never_secret(
    tmp_path, valid_config
):
    module = _preflight_module()
    workbook_path = _copy_firebase_workbook(tmp_path)
    invalid = dict(valid_config, REFLECTOR_API_KEY="")
    report = module.run_preflight(
        workbook_path=workbook_path,
        config=invalid,
        device_snapshot=_valid_snapshot(module, account_ready=None),
        expected_fingerprint=EXPECTED_FINGERPRINT,
    )
    report_path = tmp_path / "reports" / "preflight.json"

    module.write_preflight_report(report, report_path)

    serialized = report_path.read_text(encoding="utf-8")
    payload = json.loads(serialized)
    assert "super-secret-token" not in serialized
    assert '"credential_state": "configured"' in serialized
    assert '"credential_state": "missing"' in serialized
    assert payload["ready"] is False
    assert payload["blocking"] is True


class _ShellResult:
    def __init__(self, output: str, returncode: int = 0):
        self.output = output
        self.returncode = returncode


class _FakeADBDevice:
    def __init__(self, resolve_result: _ShellResult):
        self.commands: list[str] = []
        self.responses = {
            "getprop sys.boot_completed": _ShellResult("1"),
            "dumpsys power": _ShellResult("mWakefulness=Awake"),
            "dumpsys window policy": _ShellResult("mShowingLockscreen=false"),
            f"pm path com.mikirinkode.firebasechatapp": _ShellResult(
                "package:/data/app/firebase-chat/base.apk"
            ),
            (
                "cmd package resolve-activity --brief "
                "com.mikirinkode.firebasechatapp"
            ): resolve_result,
            "dumpsys activity activities": _ShellResult(
                "mResumedActivity: ActivityRecord{abc u0 "
                "com.mikirinkode.firebasechatapp/.MainActivity t1}"
            ),
            "dumpsys package com.mikirinkode.firebasechatapp": _ShellResult(""),
            "pm list packages": _ShellResult(
                "package:com.mikirinkode.firebasechatapp\npackage:com.android.settings"
            ),
        }

    def shell2(self, command: str):
        self.commands.append(command)
        return self.responses[command]


class _FakeADBClient:
    def __init__(self, device: _FakeADBDevice):
        self.device_handle = device
        self.requested_serials: list[str] = []

    def device(self, serial: str):
        self.requested_serials.append(serial)
        return self.device_handle


def test_fake_adb_client_uses_shell_queries_without_uiautomator_import(monkeypatch):
    module = _preflight_module()
    device = _FakeADBDevice(
        _ShellResult("com.mikirinkode.firebasechatapp/.MainActivity")
    )
    client = _FakeADBClient(device)
    original_import = builtins.__import__

    def reject_uiautomator(name, *args, **kwargs):
        if name == "uiautomator2" or name.startswith("uiautomator2."):
            raise AssertionError("uiautomator2 must not be used by preflight")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", reject_uiautomator)

    snapshot = module.ADBPreflightProbe(
        "fake-device-001", adb_client=client
    ).capture()

    assert client.requested_serials == ["fake-device-001"]
    assert snapshot.reachable is True
    assert snapshot.online is True
    assert snapshot.current_package == "com.mikirinkode.firebasechatapp"
    assert snapshot.current_activity == ".MainActivity"
    assert all(
        forbidden not in " ".join(device.commands).lower()
        for forbidden in (
            "app_start",
            "app_stop",
            " input ",
            " install ",
            " uninstall ",
            " pm clear",
            " pm grant",
            " pm revoke",
        )
    )


@pytest.mark.parametrize(
    ("resolve_result", "expected_launchable"),
    [
        pytest.param(
            _ShellResult("com.mikirinkode.firebasechatapp/.MainActivity", 0),
            True,
            id="valid-component",
        ),
        pytest.param(
            _ShellResult("com.mikirinkode.firebasechatapp/.MainActivity", 1),
            False,
            id="nonzero-return-code",
        ),
        pytest.param(
            _ShellResult("Error: permission denied", 0),
            False,
            id="permission-error-text",
        ),
        pytest.param(
            _ShellResult("Resolver output unavailable", 0),
            False,
            id="arbitrary-output",
        ),
        pytest.param(
            _ShellResult("com.example.other/.MainActivity", 0),
            False,
            id="wrong-package",
        ),
    ],
)
def test_launchability_requires_successful_firebase_component(
    resolve_result, expected_launchable
):
    module = _preflight_module()
    device = _FakeADBDevice(resolve_result)
    client = _FakeADBClient(device)

    snapshot = module.ADBPreflightProbe(
        "fake-device-001", adb_client=client
    ).capture()

    assert snapshot.firebase_package_launchable is expected_launchable
