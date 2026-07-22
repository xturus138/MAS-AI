import datetime
import json
import os

from core.models.state import AgentState
from core.utils.output_writer import write_run_index as _write_run_index
from core.utils.output_writer import write_run_summary as _write_run_summary
from core.utils.process_logger import LogLevel as _LL
from shared import config


class RecorderAgent:
    """
    Session finalizer — called ONCE at the end of a scenario run, not mid-cycle.

    With MIRIX, RecorderAgent is no longer a node in the LangGraph loop.
    It is called directly by the runner after the graph exits to persist final
    metrics and retrieve episodic chat history from memory.
    """

    def __init__(self, memory=None, logger=None):
        self.memory = memory
        self.logger = logger

    def _log(self, msg: str, detail: str = "", level=None):
        if self.logger is not None:
            lvl = level if level is not None else _LL.INFO
            self.logger.log("RECORDER", msg, detail, level=lvl)

    def _fill_report_sheet(
        self, dest: str, tcs_id: str, output_dir: str, metrics: dict
    ):
        import openpyxl

        wb = openpyxl.load_workbook(dest)
        ws = wb.active

        header_row_idx = None
        col_map: dict = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            first = row[0].value
            if first and str(first).strip().upper() == "TCS ID":
                header_row_idx = row_idx
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value:
                        name = str(cell.value).strip()
                        if name not in col_map:
                            col_map[name] = col_idx
                break

        if header_row_idx is None:
            print(f"[Recorder] Warning: TCS ID header row not found in {dest}")
            return

        RESULT_COLS = [
            "Time Testing",
            "Testing Status",
            "Updated At",
            "Testing By",
            "OK Evid.",
            "Issue Status",
            "Actual Result",
            "Steps Taken",
            "Tokens Used",
            "Stagnation Count",
        ]
        next_col = ws.max_column + 1
        col_index: dict = {}
        for col_name in RESULT_COLS:
            if col_name in col_map:
                col_index[col_name] = col_map[col_name]
            else:
                ws.cell(header_row_idx, next_col).value = col_name
                col_index[col_name] = next_col
                next_col += 1

        data_row_idx = None
        for row_idx in range(header_row_idx + 1, ws.max_row + 2):
            cell_val = ws.cell(row_idx, 1).value
            if cell_val and str(cell_val).strip() == tcs_id:
                data_row_idx = row_idx
                break

        if data_row_idx is None:
            print(f"[Recorder] Warning: TCS ID '{tcs_id}' not found in {dest}")
            return

        duration_s = metrics.get("total_duration_seconds", 0)
        mins, secs = divmod(int(duration_s), 60)
        duration_str = f"{mins}m {secs}s"

        status = metrics.get("status", "FAILED")
        testing_status = "OK" if status == "SUCCESS" else "NG"

        judgment = (
            metrics.get("justification", {}).get("reflector_final_judgment", "") or ""
        )

        values = {
            "Time Testing": duration_str,
            "Testing Status": testing_status,
            "Updated At": metrics.get("timestamp", ""),
            "Testing By": f"MAS AI ({metrics.get('mode', 'predefined')})",
            "OK Evid.": output_dir,
            "Issue Status": "OK" if testing_status == "OK" else judgment[:200],
            "Actual Result": judgment[:500],
            "Steps Taken": metrics.get("total_cycles", 0),
            "Tokens Used": metrics.get("total_tokens_estimate", 0),
            "Stagnation Count": metrics.get("stagnation_count", 0),
        }

        for col_name, value in values.items():
            idx = col_index.get(col_name)
            if idx:
                ws.cell(data_row_idx, idx).value = value

        wb.save(dest)
        print(f"[Recorder] Test report saved to {dest}")
        self._log(
            "Test report written",
            f"tcs_id={tcs_id}  status={testing_status}  dest={dest}",
        )

    def write_test_report(
        self, state: AgentState, metrics: dict, shared_dir: str = None
    ):
        """Generate test reports. Write to per-scenario dir, and optionally a shared run dir."""
        import shutil

        xlsx_path = os.path.join(os.getcwd(), "scenario.xlsx")
        tcs_id = state.get("tcs_id", "")
        output_dir = state.get("output_dir", "")

        if not os.path.exists(xlsx_path) or not tcs_id or not output_dir:
            return

        try:
            reports_dir = state.get("reports_dir") or os.path.join(
                output_dir, "reports"
            )
            os.makedirs(reports_dir, exist_ok=True)
            dest = os.path.join(reports_dir, "test_report.xlsx")
            shutil.copy2(xlsx_path, dest)
            self._fill_report_sheet(dest, tcs_id, output_dir, metrics)

            if shared_dir:
                shared_dest = os.path.join(shared_dir, "test_report.xlsx")
                if not os.path.exists(shared_dest):
                    shutil.copy2(xlsx_path, shared_dest)
                self._fill_report_sheet(shared_dest, tcs_id, output_dir, metrics)

        except Exception as e:
            print(f"[Recorder] Warning: could not write test report: {e}")
            self._log("Test report FAILED", str(e))

    def finalize_run_metrics(self, state: AgentState, shared_dir: str = None) -> dict:
        output_dir = state.get("output_dir", "outputs")
        logs_dir = state.get("logs_dir") or os.path.join(output_dir, "logs")
        reports_dir = state.get("reports_dir") or os.path.join(output_dir, "reports")
        os.makedirs(logs_dir, exist_ok=True)
        os.makedirs(reports_dir, exist_ok=True)
        metrics_path = os.path.join(output_dir, "final_metrics.json")
        session_id = state.get("session_id", "")
        tcs_id = state.get("tcs_id", "Unknown")

        is_completed = state.get("is_completed", False)
        stagnation = state.get("stagnation_count", 0)

        mode = "predefined"
        if self.memory is not None:
            task_goal = self.memory.core.get("task_goal") or ""
            test_type = self.memory.core.get("test_type") or ""
            sub_steps = self.memory.procedural.get_steps(tcs_id, "workflow")
            mode = "autonomous" if (task_goal and not sub_steps) else "predefined"
        else:
            mode = (
                "autonomous" if state.get("orchestrator_instruction") else "predefined"
            )

        if self.memory is not None:
            episodes = self.memory.episodic.all_as_dicts()
            chat_file = os.path.join(logs_dir, "chat_logs.txt")
            try:
                lines = []
                for ep in episodes:
                    actor = ep.get("actor", "?").upper()
                    step = ep.get("step", "?")
                    evt = ep.get("event_type", "")
                    summary = ep.get("summary", "")
                    details = ep.get("details", "")
                    lines.append(
                        f"[{actor}] (Step {step}) [{evt}]\n{summary}\n{details}\n{'-' * 40}"
                    )
                with open(chat_file, "w", encoding="utf-8") as f:
                    f.write("\n".join(lines))
            except Exception as e:
                print(f"[Recorder] Warning: could not write chat_logs.txt: {e}")

            exec_episodes = [ep for ep in episodes if ep.get("actor") == "executor"]
            script_content = {
                "tcs_id": tcs_id,
                "total_steps": len(exec_episodes),
                "recording_status": "Success" if is_completed else "In Progress/Failed",
                "actions": exec_episodes,
            }
            script_file = os.path.join(reports_dir, "interaction_script.json")
            try:
                with open(script_file, "w", encoding="utf-8") as f:
                    json.dump(script_content, f, indent=2)
            except Exception:
                pass

        total_tokens = 0
        total_cost_usd = 0.0
        if session_id:
            log_dir = state.get("llm_logs_dir") or os.path.join(
                output_dir, "logs", "llm"
            )
            if not os.path.exists(log_dir):
                legacy_log_dir = os.path.join(config.OUTPUT_DIR, "llm_logs", session_id)
                log_dir = legacy_log_dir if os.path.exists(legacy_log_dir) else log_dir
            if os.path.exists(log_dir):
                for filename in os.listdir(log_dir):
                    if filename.endswith(".json"):
                        try:
                            with open(
                                os.path.join(log_dir, filename), "r", encoding="utf-8"
                            ) as f:
                                log_data = json.load(f)
                                total_tokens += (
                                    log_data.get("total_tokens")
                                    or log_data.get("token_count_estimate")
                                    or 0
                                )
                                total_cost_usd += log_data.get("cost_usd", 0.0)
                        except Exception:
                            continue

        start_time = state.get("start_time", 0)
        end_time = state.get("end_time", 0)
        duration = end_time - start_time if end_time > start_time else 0

        status = "SUCCESS" if is_completed else "FAILED"
        if stagnation >= 3:
            status = "STAGNATED"

        exec_count = 0
        failed_count = 0
        if self.memory is not None:
            exec_eps = [
                ep
                for ep in self.memory.episodic.all_as_dicts()
                if ep.get("actor") == "executor"
            ]
            exec_count = len(exec_eps)
            failed_count = sum(
                1 for ep in exec_eps if str(ep.get("summary", "")).startswith("[FAIL]")
            )
        successful_actions = exec_count - failed_count
        tool_precision_rate = (
            round((successful_actions / exec_count) * 100, 1) if exec_count > 0 else 0.0
        )

        recovery_attempts = state.get("recovery_attempts", 0)
        if recovery_attempts > 0:
            # A "successful recovery" is a reflector pass that happened AFTER a prior failure.
            # That equals total passes minus passes on the first attempt.
            reflector_pass_count = state.get("reflector_pass_count", 0)
            first_pass_count = state.get("reflector_first_pass_count", 0)
            successful_recoveries = max(0, reflector_pass_count - first_pass_count)
            recovery_rate = round((successful_recoveries / recovery_attempts) * 100, 1)
        else:
            recovery_rate = 100.0 if is_completed else 0.0

        reflector_final_judgment = "No judgment"
        if self.memory is not None:
            last_ref = self.memory.episodic.last_by_actor("reflector")
            if last_ref:
                reflector_final_judgment = last_ref.details or last_ref.summary

        orchestrator_reasoning = ""
        if self.memory is not None:
            last_orch = self.memory.episodic.last_by_actor("orchestrator")
            if last_orch:
                orchestrator_reasoning = last_orch.summary

        figma_enabled = False
        if self.memory is not None:
            figma_enabled = (self.memory.core.get("figma_enabled") or "False") == "True"

        sub_steps_total = (
            len(self.memory.procedural.get_steps(tcs_id, "workflow"))
            if self.memory is not None
            else 0
        )
        steps_completed = state.get("steps_completed_count", 0)

        metrics = {
            "session_id": session_id,
            "tcs_id": tcs_id,
            "mode": mode,
            "status": status,
            "total_cycles": state.get("current_step", 0),
            "physical_actions": exec_count,
            "steps_completed": steps_completed,
            "stagnation_count": stagnation,
            "total_tokens_estimate": total_tokens,
            "total_price_usd": round(total_cost_usd, 8),
            "total_duration_seconds": round(duration, 2),
            "tool_precision_rate": tool_precision_rate,
            "recovery_attempts": recovery_attempts,
            "recovery_rate": recovery_rate,
            "justification": {
                "orchestrator_reasoning": orchestrator_reasoning,
                "reflector_final_judgment": reflector_final_judgment,
            },
            "figma_verified": figma_enabled,
            "last_reflector_passed": state.get("last_reflector_passed", False),
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        total_ref_calls = state.get("total_reflector_calls", 0)
        ref_passes = state.get("reflector_pass_count", 0)
        first_verify_total = state.get("total_first_verify_calls", 0)
        first_verify_passes = state.get("reflector_first_pass_count", 0)
        lookup_ok = state.get("widget_lookup_success", 0)
        lookup_fail = state.get("widget_lookup_fail", 0)
        lookup_fallback = state.get("widget_text_fallback_count", 0)

        def _pct(num, den):
            return round((num / den) * 100, 1) if den > 0 else None

        if mode == "autonomous":
            coverage_rate = None
        else:
            coverage_rate = _pct(steps_completed, sub_steps_total)

        metrics["research_metrics"] = {
            "coverage_rate": coverage_rate,
            "decision_accuracy_initial_acc1": _pct(
                first_verify_passes, first_verify_total
            ),
            "decision_accuracy_final_accf": _pct(steps_completed, first_verify_total),
            "verification_pass_rate": _pct(ref_passes, total_ref_calls),
            "widget_localization_effectiveness": _pct(
                lookup_ok, lookup_ok + lookup_fail
            ),
            "widget_text_fallback_recoveries": lookup_fallback,
            "time_overhead_seconds": round(duration, 2),
            "token_consumption": total_tokens,
        }

        try:
            with open(metrics_path, "w", encoding="utf-8") as f:
                json.dump(metrics, f, indent=4)
            print(f"[Recorder] Final metrics saved to {metrics_path}")
            self._log(
                "Final metrics saved",
                f"status={metrics['status']}  cycles={metrics['total_cycles']}\n"
                f"physical_actions={metrics['physical_actions']}\n"
                f"stagnation={metrics['stagnation_count']}\n"
                f"tokens={metrics['total_tokens_estimate']}\n"
                f"duration={metrics['total_duration_seconds']}s\n"
                f"tool_precision={metrics['tool_precision_rate']}%\n"
                f"path={metrics_path}",
            )
        except Exception as e:
            print(f"[Recorder Error] Failed to save final metrics: {e}")
            self._log("FAILED to save final metrics", str(e))

        _write_run_summary(
            output_dir=output_dir,
            tcs_id=tcs_id,
            status=status,
            steps_completed=steps_completed,
            duration_seconds=duration,
        )

        self.write_test_report(state, metrics, shared_dir=shared_dir)

        if self.logger is not None:
            self.logger.close()

        return metrics

    @staticmethod
    def write_run_summary(all_metrics: list, shared_dir: str):
        """Aggregate per-scenario metrics into a single run-level summary."""
        if not all_metrics or not shared_dir:
            return

        total_tokens = sum(m.get("total_tokens_estimate", 0) for m in all_metrics)
        total_cost = sum(m.get("total_price_usd", 0.0) for m in all_metrics)
        total_duration = sum(m.get("total_duration_seconds", 0.0) for m in all_metrics)
        total_scenarios = len(all_metrics)
        passed_scenarios = sum(1 for m in all_metrics if m.get("status") == "SUCCESS")
        overall_pass_rate = (
            round((passed_scenarios / total_scenarios) * 100, 1)
            if total_scenarios > 0
            else 0.0
        )

        total_physical_actions = sum(m.get("physical_actions", 0) for m in all_metrics)
        total_recovery_attempts = sum(
            m.get("recovery_attempts", 0) for m in all_metrics
        )

        cov_rates = [
            m.get("research_metrics", {}).get("coverage_rate")
            for m in all_metrics
            if m.get("research_metrics", {}).get("coverage_rate") is not None
        ]
        avg_coverage = round(sum(cov_rates) / len(cov_rates), 1) if cov_rates else None

        tool_precisions = [m.get("tool_precision_rate", 0) for m in all_metrics]
        avg_tool_precision = (
            round(sum(tool_precisions) / len(tool_precisions), 1)
            if tool_precisions
            else 0.0
        )

        per_scenario_status = []
        for m in all_metrics:
            per_scenario_status.append(
                {
                    "tcs_id": m.get("tcs_id"),
                    "status": m.get("status"),
                    "tokens": m.get("total_tokens_estimate"),
                    "duration_seconds": m.get("total_duration_seconds"),
                    "coverage_rate": m.get("research_metrics", {}).get("coverage_rate"),
                }
            )

        summary = {
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_scenarios": total_scenarios,
            "passed_scenarios": passed_scenarios,
            "overall_pass_rate": overall_pass_rate,
            "total_tokens": total_tokens,
            "total_cost_usd": round(total_cost, 8),
            "total_duration_seconds": round(total_duration, 2),
            "total_physical_actions": total_physical_actions,
            "total_recovery_attempts": total_recovery_attempts,
            "avg_coverage_rate": avg_coverage,
            "avg_tool_precision": avg_tool_precision,
            "scenario_details": per_scenario_status,
        }

        summary_path = os.path.join(shared_dir, "run_summary.json")
        try:
            with open(summary_path, "w", encoding="utf-8") as f:
                json.dump(summary, f, indent=4)
            print(f"[Recorder] Merged run summary saved to {summary_path}")
        except Exception as e:
            print(f"[Recorder Error] Failed to save run summary: {e}")

        _write_run_index(shared_dir, all_metrics)
